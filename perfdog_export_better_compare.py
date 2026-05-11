import os
import ntpath
import sys
import argparse
import statistics
import re
import math
from datetime import datetime
try:
    import pandas as pd
except ImportError:
    os.system('pip install pandas')
    import pandas as pd
try:
    import json
except ImportError:
    os.system('pip install json')
    import json
try:
    import openpyxl
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.formatting.rule import DataBarRule
except ImportError:
    os.system('pip install openpyxl')
    import openpyxl
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.formatting.rule import DataBarRule
try:
    import numpy as np
except ImportError:
    os.system('pip install numpy')
    import numpy as np
try:
    import matplotlib as mpl
except ImportError:
    os.system('pip install matplotlib')
    import matplotlib as mpl




def process_data(input_data_list, input_perfdog_config, output_xlsx, divided_by_framerate, divided_by_resolution, compare_target_column_name, compare_target_name, show_only_columns_in_config, sort_vs_by_value=True):

    input_data_list = [os.path.normpath(path) for path in input_data_list]
    if not output_xlsx:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_xlsx = os.path.splitext(input_data_list[0])[0] + f"_better_compare_{timestamp}.xlsx"

    print("=" * 60)
    print("  PerfDog Better Compare")
    print("=" * 60)
    print(f"  Input  : {input_data_list}")
    print(f"  Output : {output_xlsx}")
    print(f"  Normalize by framerate  : {divided_by_framerate}")
    print(f"  Normalize by resolution : {divided_by_resolution}")
    print(f"  Compare column : {compare_target_column_name or '(auto)'}")
    print(f"  Compare target : {compare_target_name or '(auto)'}")
    print(f"  Show only config columns: {show_only_columns_in_config}")
    print(f"  Sort VS sheet  : {sort_vs_by_value}")

    if not input_perfdog_config:
        current_script_path = os.path.abspath(__file__)
        current_script_dir = os.path.dirname(current_script_path)
        input_perfdog_config = current_script_dir + "/perfdog_export_better_compare_config.json"
        print(f"  Config : (default) {input_perfdog_config}")
    input_perfdog_config = os.path.normpath(input_perfdog_config)
    print(f"  Config : {input_perfdog_config}")
    print("-" * 60)

    try:
        # read the json config
        with open(input_perfdog_config, encoding='utf-8') as json_file:
            config_data = json.load(json_file)
    except FileNotFoundError:
        print(f"Error: File '{input_perfdog_config}' not found.")
        sys.exit(1)
    except IOError:
        print(f"Error: Unable to open file '{input_perfdog_config}'.")
        sys.exit(1)

    if not compare_target_column_name:
        compare_target_column_name = config_data["test_case_column"]
        print(f"  [auto] compare_target_column_name -> '{compare_target_column_name}'")

    # save original value; resolve after data is loaded
    _auto_compare_target_name = compare_target_name

    # convert kilo/mega/giga prefixes to numeric values
    def convert_unit_prefix(cell_value):
        if 'kilo' in str(cell_value).lower():
            return float(cell_value.lower().replace('kilo', '')) * 1e3
        elif 'mega' in str(cell_value).lower():
            return float(cell_value.lower().replace('mega', '')) * 1e6
        elif 'giga' in str(cell_value).lower():
            return float(cell_value.lower().replace('giga', '')) * 1e9
        else:
            return cell_value

    input_df_list = list()
    for one_input_data_name in input_data_list:
        temp_df = pd.read_excel(one_input_data_name)
        temp_df.columns = temp_df.columns.str.replace('\n', ' ')
        for column in temp_df.columns:
            temp_df[column] = temp_df[column].apply(convert_unit_prefix)
        input_df_list.append(temp_df)

    # merge all input files
    df = input_df_list[0].copy()
    for input_df in input_df_list[1:]:
        df = pd.concat([df, input_df], ignore_index=True)

    # aggregate: mean for numeric columns, newline-join for text columns
    def custom_agg(x):
        if pd.to_numeric(x, errors='coerce').notna().all():
            return pd.to_numeric(x, errors='coerce').mean()
        else:
            return x.astype(str).str.cat(sep='\n')

    df = df.groupby(compare_target_column_name).agg(custom_agg).reset_index()

    # default compare target: first row in the grouped data
    if not _auto_compare_target_name and compare_target_column_name in df.columns and len(df) > 1:
        _auto_compare_target_name = df[compare_target_column_name].iloc[0]
        print(f"  [auto] compare_target_name -> '{_auto_compare_target_name}'")
    compare_target_name = _auto_compare_target_name

    # prompt user for missing important columns
    important_columns = [config_data["average_framerate_column"]]
    for column in important_columns:
        if column not in df.columns:
            print(f"  [warn] Missing important column: '{column}'")
            for index, row in df.iterrows():
                value = input(f"  Please enter value for test case '{row[compare_target_column_name]}' / column '{column}': ")
                df.at[index, column] = value

    resolution_columns = [config_data["resolution_width_column"], config_data["resolution_height_column"]]
    for column in resolution_columns:
        if column not in df.columns:
            if divided_by_resolution:
                print(f"  [warn] Missing resolution column: '{column}'")
            else:
                print(f"  [info] Missing resolution column: '{column}', reset to 1")
            for index, row in df.iterrows():
                if divided_by_resolution:
                    value = input(f"  Please enter value for test case '{row[compare_target_column_name]}' / column '{column}': ")
                    df.at[index, column] = value
                else:
                    df.at[index, column] = 1


    # make the normalized data sheet copy
    df_processed = df.copy()

    # only show specified columns
    if show_only_columns_in_config:
        all_config_columns = [config_data["average_framerate_column"], config_data["resolution_height_column"], config_data["resolution_width_column"]]
        all_config_columns += config_data['columns_always_shown'] + config_data['columns_divide_by_framerate'] + config_data['columns_divide_by_resolution'] + config_data['columns_divide_by_framerate_and_resolution']
        columns_to_keep = [column for column in df_processed.columns if any(config_col in column for config_col in all_config_columns)]
        df_processed = df_processed[columns_to_keep]

    # processing the data
    for column in df_processed.columns:
        if any(config_col in column for config_col in config_data['columns_divide_by_framerate']):
            if divided_by_framerate:
                df_processed[column] = df_processed[column].astype(float) / df_processed[config_data["average_framerate_column"]].astype(float)
                df_processed.rename(columns={column: column + '\n(/framerate)'}, inplace=True)
        elif any(config_col in column for config_col in config_data['columns_divide_by_resolution']):
            if divided_by_resolution:
                df_processed[column] = df_processed[column].astype(float) / (df_processed[config_data["resolution_height_column"]].astype(float) * df_processed[config_data["resolution_width_column"]].astype(float))
                df_processed.rename(columns={column: column + '\n(/resolution)'}, inplace=True)
        elif any(config_col in column for config_col in config_data['columns_divide_by_framerate_and_resolution']):
            if divided_by_framerate and divided_by_resolution:
                df_processed[column] = df_processed[column].astype(float) / (df_processed[config_data["average_framerate_column"]].astype(float) * df_processed[config_data["resolution_height_column"]].astype(float) * df_processed[config_data["resolution_width_column"]].astype(float))
                df_processed.rename(columns={column: column + '\n(/(framerate*resolution))'}, inplace=True)


    def fix_filename(filename):
        # sanitize illegal characters (Windows)
        illegal_chars = r'[<>:："/\\|?*]'
        filename = re.sub(illegal_chars, '_', filename)
        # truncate if adding " VS. Others" would exceed 31 chars (Excel sheet name limit)
        max_name_length = 31 - len(" VS. Others")
        if len(filename) > max_name_length:
            filename = filename[:max_name_length]
        return filename

    df_compare_target = None
    raw_values_dict = {}
    target_project_compare_sheet_name = ""
    if compare_target_column_name and compare_target_name:
        # find the target row
        target_row = df_processed.loc[df_processed[compare_target_column_name] == compare_target_name]
        if not target_row.empty:
            target_project_compare_sheet_name = fix_filename(compare_target_name) + " VS. Others"
            comparison_rows = []
            raw_values_list = []  # parallel list to comparison_rows

            for index, row in df_processed.iterrows():
                if row[compare_target_column_name] == compare_target_name:
                    continue

                comparison_row = {compare_target_column_name: f"{compare_target_name}\n VS. \n{row[compare_target_column_name]}"}
                row_raw_values = {}

                for col in df_processed.columns[1:]:
                    target_value = target_row[col].values[0]
                    other_value = row[col]

                    if isinstance(target_value, (int, float)) and isinstance(other_value, (int, float)) and not np.isclose(other_value, 0.0):
                        comparison_row[col] = target_value / other_value
                        row_raw_values[col] = f"{target_value} / {other_value}"
                    else:
                        comparison_row[col] = f"{target_value} / {other_value}"
                        row_raw_values[col] = f"{target_value} / {other_value}"

                comparison_rows.append(comparison_row)
                raw_values_list.append(row_raw_values)

            # build DataFrame at once to avoid fragmentation from row-by-row loc appends
            df_compare_target = pd.DataFrame(comparison_rows, columns=df_processed.columns)
            raw_values_dict = {i: v for i, v in enumerate(raw_values_list)}

            # sort columns by ratio (desc): fixed header cols stay first, non-numeric cols sink to bottom
            if sort_vs_by_value and not df_compare_target.empty:
                all_cols = list(df_compare_target.columns)
                frozen_n = config_data.get('forzen_column_num', 3)
                fixed_cols = all_cols[:frozen_n]
                metric_cols = all_cols[frozen_n:]

                def col_sort_key(col_name):
                    vals = df_compare_target[col_name]
                    numeric_vals = [v for v in vals if isinstance(v, (int, float))]
                    return statistics.mean(numeric_vals) if numeric_vals else float('-inf')

                numeric_metric_cols = sorted(
                    [c for c in metric_cols if col_sort_key(c) > float('-inf')],
                    key=col_sort_key, reverse=True
                )
                non_numeric_metric_cols = [c for c in metric_cols if col_sort_key(c) == float('-inf')]
                sorted_cols = fixed_cols + numeric_metric_cols + non_numeric_metric_cols
                df_compare_target = df_compare_target[sorted_cols]

            print(f"  [info] VS sheet: '{target_project_compare_sheet_name}'")
        else:
            print(f"  [warn] Cannot find compare target '{compare_target_name}', VS sheet will not be generated.")


    def path_leaf(path):
        head, tail = ntpath.split(path)
        return tail or ntpath.basename(head)
        
    # 生成临时的标准格式文件
    base_output_xlsx = os.path.splitext(output_xlsx)[0] + "_base.xlsx"
    
    # write to intermediate base file (VS sheet first if present, then BarCompare, then CompareSource)
    with pd.ExcelWriter(base_output_xlsx) as writer:
        if df_compare_target is not None:
            df_compare_target.to_excel(writer, sheet_name=target_project_compare_sheet_name, index=False)
        df_processed.to_excel(writer, sheet_name='BarCompare', index=False)
        # CompareSource sheets keep raw data as-is
        for i, one_df in enumerate(input_df_list):
            one_df.to_excel(writer, sheet_name='CompareSource' + str(i), index=False)

    # apply styles to the intermediate file
    wb = openpyxl.load_workbook(base_output_xlsx)

    # apply data bars to BarCompare sheet
    ws = wb['BarCompare']
    for i, column in enumerate(ws.columns):
        column = [cell for cell in column]
        if all(isinstance(cell.value, (int, float)) for cell in column[1:]):
            values = [cell.value for cell in column[1:]]
            if values:
                max_value = max(values)
                average_value = statistics.mean(values)
                rule = DataBarRule(start_type="num", start_value=max_value - average_value, end_type="num", end_value=max_value, color="999999")
                ws.conditional_formatting.add(openpyxl.utils.get_column_letter(i + 1) + '2:' + openpyxl.utils.get_column_letter(i + 1) + str(ws.max_row), rule)

    # apply styles and comments to VS sheet
    if df_compare_target is not None:
        ws_target_project_compare_sheet = wb[target_project_compare_sheet_name]

        for i, row in enumerate(ws_target_project_compare_sheet.iter_rows(), 1):
            values = []
            for j, cell in enumerate(row, 1):
                if isinstance(cell.value, (int, float)):
                    values.append(cell.value)
                    # add raw value comment
                    try:
                        row_idx = i - 2  # skip header row; dict keys start at 0
                        if row_idx >= 0 and j < len(df_compare_target.columns):
                            col_name = df_compare_target.columns[j]
                            if row_idx in raw_values_dict and col_name in raw_values_dict[row_idx]:
                                raw_value = raw_values_dict[row_idx][col_name]
                                if raw_value:
                                    cell.comment = openpyxl.comments.Comment(
                                        raw_value,
                                        'PerfDog Better Compare'
                                    )
                    except Exception as e:
                        print(f"  [warn] Failed to add comment at row {i}, col {j}: {e}")
                        continue

            if len(values) > 0:
                q1 = np.percentile(values, 25)
                q3 = np.percentile(values, 75)
                iqr = q3 - q1
                lower_bound = q1 - (1.5 * iqr)
                upper_bound = q3 + (1.5 * iqr)

                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.fill = color_cell(cell.value, lower_bound, upper_bound)
                        cell.number_format = "0.00%"

    # apply header styles to all non-CompareSource sheets
    important_background_fill = openpyxl.styles.PatternFill(start_color="f9d3e3", end_color="f9d3e3", fill_type="solid")
    for one_ws_name in wb.sheetnames:
        if one_ws_name.startswith('CompareSource'):
            continue

        one_ws = wb[one_ws_name]
        # wrap text and highlight important columns in header row
        for cell in one_ws[1]:
            cell.alignment = Alignment(wrap_text=True)
            if cell.value is not None:
                if any(config_col in str(cell.value) for config_col in config_data['columns_important_background']):
                    cell.fill = important_background_fill

        # auto row height based on line count in header
        max_lines = max(
            (str(cell.value).count('\n') + 1 for cell in one_ws[1] if cell.value is not None),
            default=1
        )
        one_ws.row_dimensions[1].height = max_lines * 80

        # set default column width
        for column in one_ws.columns:
            one_ws.column_dimensions[column[0].column_letter].width = 15

    wb.save(base_output_xlsx)

    # build transposed output workbook
    wb_transposed = openpyxl.Workbook()
    wb = openpyxl.load_workbook(base_output_xlsx)

    # VS sheet first, then BarCompare
    sheets_to_transpose = []
    if df_compare_target is not None:
        sheets_to_transpose.append(target_project_compare_sheet_name)
    sheets_to_transpose.append('BarCompare')

    for sheet_name in sheets_to_transpose:
        if sheet_name in wb.sheetnames:
            ws_source = wb[sheet_name]

            if sheet_name in wb_transposed.sheetnames:
                wb_transposed.remove(wb_transposed[sheet_name])

            ws_target = wb_transposed.create_sheet(sheet_name)

            # read and transpose data
            data = [[cell.value for cell in row] for row in ws_source.rows]
            transposed_data = list(zip(*data))

            # write transposed data with formatting
            for i, row in enumerate(transposed_data, 1):
                values = []  # numeric values in this row, for data bar / heatmap
                for j, value in enumerate(row, 1):
                    cell = ws_target.cell(row=i, column=j, value=value)

                    if i == 1:  # header row (first col after transpose)
                        cell.alignment = Alignment(wrap_text=True)
                    elif sheet_name == target_project_compare_sheet_name:
                        if isinstance(value, (int, float)):
                            values.append(value)
                            cell.number_format = "0.00%"
                            # after transpose: row i → original column i-1; col j → original row j-2
                            try:
                                if j >= 2 and i - 1 < len(df_compare_target.columns):
                                    original_col = df_compare_target.columns[i-1]
                                    row_idx = j - 2
                                    if row_idx in raw_values_dict and original_col in raw_values_dict[row_idx]:
                                        raw_value = raw_values_dict[row_idx][original_col]
                                        if raw_value:
                                            cell.comment = openpyxl.comments.Comment(
                                                text=raw_value,
                                                author='PerfDog Better Compare'
                                            )
                            except Exception as e:
                                print(f"  [warn] Failed to add comment at row {i}, col {j}: {e}")

                    elif sheet_name == 'BarCompare':
                        if isinstance(value, (int, float)):
                            values.append(value)

                # apply per-row formatting after collecting all values
                if i > 1:
                    if sheet_name == 'BarCompare' and values:
                        max_value = max(values)
                        average_value = statistics.mean(values)
                        rule = DataBarRule(
                            start_type="num", start_value=max_value - average_value,
                            end_type="num", end_value=max_value,
                            color="999999"
                        )
                        ws_target.conditional_formatting.add(
                            f"B{i}:{openpyxl.utils.get_column_letter(ws_target.max_column)}{i}",
                            rule
                        )
                    elif sheet_name == target_project_compare_sheet_name and values:
                        q1 = np.percentile(values, 25)
                        q3 = np.percentile(values, 75)
                        iqr = q3 - q1
                        lower_bound = q1 - (1.5 * iqr)
                        upper_bound = q3 + (1.5 * iqr)
                        for j, value in enumerate(row, 1):
                            if isinstance(value, (int, float)):
                                cell = ws_target.cell(row=i, column=j)
                                cell.fill = color_cell(value, lower_bound, upper_bound)

                # handle first column (metric name) formatting
                if i > 1:
                    first_cell = ws_target.cell(row=i, column=1)
                    if first_cell.value is not None:
                        if str(first_cell.value) in [config_data["project_column"], config_data["test_case_column"]]:
                            first_cell.alignment = Alignment(wrap_text=True)
                            if '\n' in str(first_cell.value):
                                ws_target.row_dimensions[i].height = (str(first_cell.value).count('\n') + 1) * 80
                        else:
                            first_cell.alignment = Alignment(wrap_text=False)
                            # estimate display width: CJK chars count as 2, ASCII as 1
                            text = str(first_cell.value)
                            char_width = sum(2 if ord(c) > 127 else 1 for c in text)
                            current_width = ws_target.column_dimensions['A'].width if ws_target.column_dimensions['A'].width else 15
                            ws_target.column_dimensions['A'].width = max(current_width, char_width + 2)

                    # left-align data cells in VS sheet
                    if sheet_name == target_project_compare_sheet_name and i == 1:
                        for j in range(2, len(row) + 1):
                            c = ws_target.cell(row=i, column=j)
                            c.alignment = Alignment(horizontal='left', wrap_text=True)

            # apply left alignment to all data cells (col >= 2) in VS sheet header row (row 1)
            if sheet_name == target_project_compare_sheet_name:
                for j in range(2, ws_target.max_column + 1):
                    c = ws_target.cell(row=1, column=j)
                    c.alignment = Alignment(horizontal='left', wrap_text=True)

            # data columns: fixed width 22
            for col in range(2, ws_target.max_column + 1):
                ws_target.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

            # highlight important metric rows
            important_background_fill = openpyxl.styles.PatternFill(start_color="f9d3e3", end_color="f9d3e3", fill_type="solid")
            for i, row in enumerate(transposed_data, 1):
                cell = ws_target.cell(row=i, column=1)
                if cell.value is not None:
                    cell_str = str(cell.value)
                    base_name = cell_str.split('\n')[0] if '\n' in cell_str else cell_str
                    for config_col in config_data['columns_important_background']:
                        if config_col in base_name:
                            cell.fill = important_background_fill
                            break

            # freeze panes: freeze first column + first (forzen_column_num) rows
            ws_target.freeze_panes = ws_target.cell(row=config_data.get('forzen_column_num', 3) + 1, column=2)

    # copy CompareSource sheets as-is (no transpose)
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith('CompareSource'):
            ws_source = wb[sheet_name]
            ws_target = wb_transposed.create_sheet(sheet_name)

            for row in ws_source.rows:
                for cell in row:
                    new_cell = ws_target.cell(row=cell.row, column=cell.column, value=cell.value)
                    # copy cell format
                    if cell.alignment:
                        new_cell.alignment = Alignment(
                            horizontal=cell.alignment.horizontal,
                            vertical=cell.alignment.vertical,
                            text_rotation=cell.alignment.text_rotation,
                            wrap_text=cell.alignment.wrap_text,
                            shrink_to_fit=cell.alignment.shrink_to_fit,
                            indent=cell.alignment.indent
                        )
                    if cell.fill:
                        new_cell.fill = PatternFill(
                            start_color=cell.fill.start_color.rgb,
                            end_color=cell.fill.end_color.rgb,
                            fill_type=cell.fill.fill_type
                        )
                    if cell.comment:
                        new_cell.comment = openpyxl.comments.Comment(
                            text=cell.comment.text,
                            author=cell.comment.author
                        )
            
            # copy column widths
            for column in ws_source.columns:
                ws_target.column_dimensions[column[0].column_letter].width = ws_source.column_dimensions[column[0].column_letter].width

            # copy row heights
            for row in ws_source.row_dimensions:
                if row in ws_source.row_dimensions:
                    ws_target.row_dimensions[row] = ws_source.row_dimensions[row]

    # remove default empty sheet created by openpyxl
    if 'Sheet' in wb_transposed.sheetnames:
        wb_transposed.remove(wb_transposed['Sheet'])
    
    # save final output
    wb_transposed.save(output_xlsx)

    # remove intermediate file
    os.remove(base_output_xlsx)

    print(f"  [done] Output saved: {output_xlsx}")
    print("=" * 60)

def interpolate_among_3color(percentage, color0, color1, color2):
    # build piecewise linear color map
    cmap_dict = {
        'red':   [(0.0, color0[0], color0[0]), (0.5, color1[0], color1[0]), (1.0, color2[0], color2[0])],
        'green': [(0.0, color0[1], color0[1]), (0.5, color1[1], color1[1]), (1.0, color2[1], color2[1])],
        'blue':  [(0.0, color0[2], color0[2]), (0.5, color1[2], color1[2]), (1.0, color2[2], color2[2])],
    }
    cmap = mpl.colors.LinearSegmentedColormap('custom_cmap', cmap_dict)
    return cmap(percentage)

def rgb_to_hex(rgb):
    return "{:02x}{:02x}{:02x}".format(int(rgb[0] * 255), int(rgb[1] * 255), int(rgb[2] * 255))

NEUTRAL_ZONE_LOW  = 0.90  # below this: pure blue gradient
NEUTRAL_ZONE_HIGH = 1.10  # above this: orange-red gradient
COLOR_NEUTRAL = (0.75, 0.75, 0.75)  # gray at 100%

def color_cell(value, min_value, max_value):
    """
    Smooth gradient using log scale centered at 1.0 (100%):
      value << 1  → blue
      value ≈ 0.9 → blue-gray blend
      value = 1.0 → gray
      value ≈ 1.1 → gray-orange blend
      value >> 1  → red
    No hard cutoffs; the ±10% neutral zone fades naturally through gray.
    """
    color_blue   = (0, 0.44, 0.75)
    color_gray   = (0.75, 0.75, 0.75)
    color_orange = (1, 0.75, 0)
    color_red    = (1, 0, 0)

    def make_fill(rgb):
        h = rgb_to_hex(rgb)
        return PatternFill(start_color=h, end_color=h, fill_type="solid")

    # map value to [0,1] using log scale; log(1.0)=0 maps to 0.5 (gray center)
    # ±log(1.1) ≈ ±0.095 maps to the ±10% neutral boundary
    # use log(2) ≈ 0.693 as the saturation scale so that 200% → ~1.0 and 50% → ~0.0
    log_val = math.log(max(value, 1e-9))
    scale = math.log(2)  # half-saturation at 2x or 0.5x
    ratio = 0.5 + log_val / (2 * scale)
    ratio = max(0.0, min(1.0, ratio))

    # 5-stop color map: blue → blue → gray → orange → red
    # stops at [0.0, 0.35, 0.5, 0.65, 1.0]
    if ratio <= 0.35:
        t = ratio / 0.35
        rgb = interpolate_among_3color(t, color_blue, color_blue, color_blue)
    elif ratio <= 0.5:
        t = (ratio - 0.35) / (0.5 - 0.35)
        rgb = interpolate_among_3color(t, color_blue, color_gray, color_gray)
    elif ratio <= 0.65:
        t = (ratio - 0.5) / (0.65 - 0.5)
        rgb = interpolate_among_3color(t, color_gray, color_gray, color_orange)
    else:
        t = (ratio - 0.65) / (1.0 - 0.65)
        rgb = interpolate_among_3color(t, color_orange, color_orange, color_red)

    return make_fill(rgb)

def main():
    parser = argparse.ArgumentParser(description='Process PerfDog exported Excel files and generate better comparison.')
    
    # 添加位置参数，用于单个文件输入的情况
    parser.add_argument('input_file', nargs='?', help='input a single PerfDog exported xlsx file')
    
    # 原有的列表参数，现在变为可选
    parser.add_argument('-i', '--input_data_list', nargs='+', help='input multiple PerfDog exported xlsx files. multiple xlsx stats will be averaged for each project.')
    parser.add_argument('-c', '--input_perfdog_config', help='Input PerfDog config file path (json format). You can specify the provided perfdog_export_better_compare_config.json')
    parser.add_argument('-o', '--output_xlsx', help='output file path. if not specified, OUTPUT_XLSX will be INPUT_DATA_LIST[0]_better_compare.xlsx')
    
    parser.add_argument('-f', '--divided_by_framerate', action='store_true', help='false by default, whether normalized some columns by the framerate, see also perfdog_export_better_compare_config.json')
    parser.add_argument('-r', '--divided_by_resolution', action='store_true', help='false by default, whether normalized some columns by the resolution, see also perfdog_export_better_compare_config.json')

    parser.add_argument('-C', '--compare_target_column_name', help='optional, default to 用例, you may change to 项目')
    parser.add_argument('-t', '--compare_target_name', help='optional, input one target name and generate the "Target VS. Others" sheet. target name is one of values in compare_target_column_name column (default is 用例, you may change to 项目 by the --compare_target_column_name param')
    
    parser.add_argument('-s', '--show_only_columns_in_config', help='Normalized sheet show only columns in config json', default=False)
    parser.add_argument('-n', '--no_sort_vs', action='store_true', help='disable descending sort by ratio value in the VS. sheet (sort is enabled by default)')

    args = parser.parse_args()

    # 处理输入文件参数
    input_data_list = []
    if args.input_file:
        input_data_list = [args.input_file]
    elif args.input_data_list:
        input_data_list = args.input_data_list
    else:
        parser.error('No input files specified. Please provide either a single input file or use --input_data_list for multiple files.')

    process_data(input_data_list, args.input_perfdog_config, args.output_xlsx, args.divided_by_framerate, args.divided_by_resolution, args.compare_target_column_name, args.compare_target_name, args.show_only_columns_in_config, sort_vs_by_value=not args.no_sort_vs)

if __name__ == '__main__':
    main()